import os
import pytest
from openpyxl import Workbook
from excel_converter import (
    excel_to_json,
    extract_cell_dependencies,
    process_array_formula,
    get_column_letter
)

@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Add some basic data
    ws['A1'] = "Header"
    ws['A2'] = 1
    ws['B2'] = 2
    ws['C2'] = "=SUM(A2:B2)"
    
    # Add an array formula
    ws['A4'] = "{=SUM(A2:B2)}"
    
    # Save the workbook
    test_file = "test_workbook.xlsx"
    wb.save(test_file)
    
    yield test_file
    
    # Cleanup
    if os.path.exists(test_file):
        os.remove(test_file)

def test_excel_to_json_basic(sample_excel_file):
    """Test basic Excel to JSON conversion."""
    result = excel_to_json(sample_excel_file)
    
    assert result is not None
    assert "fileName" in result
    assert "sheets" in result
    assert len(result["sheets"]) == 1
    
    sheet = result["sheets"][0]
    assert sheet["sheetTitle"] == "Test Sheet"
    assert "A1" in sheet["cells"]
    assert sheet["cells"]["A1"]["value"] == "Header"
    assert "C2" in sheet["cells"]
    assert sheet["cells"]["C2"]["formula"] == "=SUM(A2:B2)"

def test_extract_cell_dependencies():
    """Test cell dependency extraction from formulas."""
    formula = "=SUM(A1:B10) + C5"
    deps = extract_cell_dependencies(formula)
    
    assert "A1:B10" in deps
    assert "C5" in deps
    assert len(deps) == 2

def test_process_array_formula():
    """Test array formula processing."""
    class MockCell:
        def __init__(self):
            self.value = "{=SUM(A1:B1)}"
            self.coordinate = "C1"
            self.internal_value = 3

    cell = MockCell()
    result = process_array_formula(cell)
    
    assert result is not None
    assert result["type"] == "array_formula"
    assert result["formula"] == "=SUM(A1:B1)"
    assert result["range"] == "C1"

def test_get_column_letter():
    """Test column number to letter conversion."""
    assert get_column_letter(1) == "A"
    assert get_column_letter(26) == "Z"
    assert get_column_letter(27) == "AA"
    assert get_column_letter(52) == "AZ"

def test_excel_to_json_with_sample_size(sample_excel_file):
    """Test Excel to JSON conversion with sample size limit."""
    result = excel_to_json(sample_excel_file, sample_size=2)
    
    assert result is not None
    assert len(result["sheets"]) == 1
    sheet = result["sheets"][0]
    
    # Should only include cells from first two rows
    assert "A1" in sheet["cells"]
    assert "A2" in sheet["cells"]
    assert "A4" not in sheet["cells"]  # This is in row 4, should be excluded

def test_excel_to_json_invalid_file():
    """Test Excel to JSON conversion with invalid file."""
    result = excel_to_json("nonexistent.xlsx")
    assert result is None

def test_real_spreadsheet():
    """Test conversion of the actual testspreadsheet.xlsx file."""
    result = excel_to_json("testspreadsheet.xlsx")
    
    assert result is not None
    assert result["fileName"] == "testspreadsheet.xlsx"
    assert "sheets" in result
    
    # Test sheet names
    sheet_names = {sheet["sheetTitle"] for sheet in result["sheets"]}
    expected_sheets = {"SAF", "SAF Weather", "WM", "WM Weather", "YM", "YM Weather"}
    assert sheet_names == expected_sheets
    
    # Test sheet structure
    for sheet in result["sheets"]:
        assert "maxRow" in sheet
        assert "maxColumn" in sheet
        assert "cells" in sheet
        assert isinstance(sheet["cells"], dict)
        
        # Test cell structure
        for coord, cell in sheet["cells"].items():
            assert "value" in cell
            assert "style" in cell
            assert isinstance(cell["style"], dict)
            assert "font" in cell["style"]
            assert "fill" in cell["style"]
            assert "alignment" in cell["style"]

def test_real_spreadsheet_with_sampling():
    """Test sampling functionality with the actual spreadsheet."""
    sample_size = 5
    result = excel_to_json("testspreadsheet.xlsx", sample_size=sample_size)
    
    assert result is not None
    
    # Check that each sheet respects the sample size
    for sheet in result["sheets"]:
        row_numbers = {int(coord[1:]) for coord in sheet["cells"].keys() if coord[1:].isdigit()}
        assert max(row_numbers) <= sample_size

@pytest.mark.parametrize("sample_size", [1, 5, 10])
def test_real_spreadsheet_various_samples(sample_size):
    """Test different sample sizes with the actual spreadsheet."""
    result = excel_to_json("testspreadsheet.xlsx", sample_size=sample_size)
    
    assert result is not None
    
    for sheet in result["sheets"]:
        row_numbers = {int(coord[1:]) for coord in sheet["cells"].keys() if coord[1:].isdigit()}
        if row_numbers:  # If there are any numbered rows
            assert max(row_numbers) <= sample_size

def test_formula_extraction_real():
    """Test formula extraction from the actual spreadsheet."""
    result = excel_to_json("testspreadsheet.xlsx")
    
    for sheet in result["sheets"]:
        for cell_data in sheet["cells"].values():
            if "formula" in cell_data:
                assert isinstance(cell_data["formula"], str)
                assert cell_data["formula"].startswith("=")
                if "dependencies" in cell_data:
                    assert isinstance(cell_data["dependencies"], list)

def test_style_information_real():
    """Test style information extraction from the actual spreadsheet."""
    result = excel_to_json("testspreadsheet.xlsx")
    
    for sheet in result["sheets"]:
        for cell_data in sheet["cells"].values():
            style = cell_data["style"]
            
            # Font checks
            assert "bold" in style["font"]
            assert "italic" in style["font"]
            assert "color" in style["font"]
            
            # Fill checks
            assert "background" in style["fill"]
            
            # Alignment checks
            assert "horizontal" in style["alignment"]
            assert "vertical" in style["alignment"]

def test_spreadsheet2():
    """Test conversion of testspreadshee2.xlsx file."""
    result = excel_to_json("testspreadshee2.xlsx")
    
    assert result is not None
    assert result["fileName"] == "testspreadshee2.xlsx"
    assert "sheets" in result
    
    # Test sheet structure
    for sheet in result["sheets"]:
        assert "maxRow" in sheet
        assert "maxColumn" in sheet
        assert "cells" in sheet
        assert isinstance(sheet["cells"], dict)
        
        # Test cell structure
        for coord, cell in sheet["cells"].items():
            assert "value" in cell
            assert "style" in cell
            assert isinstance(cell["style"], dict)
            assert "font" in cell["style"]
            assert "fill" in cell["style"]
            assert "alignment" in cell["style"]

def test_spreadsheet2_with_sampling():
    """Test sampling functionality with testspreadshee2.xlsx."""
    sample_size = 5
    result = excel_to_json("testspreadshee2.xlsx", sample_size=sample_size)
    
    assert result is not None
    
    # Check that each sheet respects the sample size
    for sheet in result["sheets"]:
        row_numbers = {int(coord[1:]) for coord in sheet["cells"].keys() if coord[1:].isdigit()}
        if row_numbers:  # If there are any numbered rows
            assert max(row_numbers) <= sample_size

def test_spreadsheet2_formulas():
    """Test formula extraction from testspreadshee2.xlsx."""
    result = excel_to_json("testspreadshee2.xlsx")
    
    formula_count = 0
    for sheet in result["sheets"]:
        for cell_data in sheet["cells"].values():
            if "formula" in cell_data:
                formula_count += 1
                assert isinstance(cell_data["formula"], str)
                assert cell_data["formula"].startswith("=")
                if "dependencies" in cell_data:
                    assert isinstance(cell_data["dependencies"], list)
    
    print(f"Found {formula_count} formulas in testspreadshee2.xlsx")

def test_spreadsheet2_cell_values():
    """Test specific cell values in testspreadshee2.xlsx."""
    result = excel_to_json("testspreadshee2.xlsx")
    
    assert result is not None
    
    # Print some statistics about the data
    for sheet in result["sheets"]:
        print(f"\nSheet: {sheet['sheetTitle']}")
        print(f"Dimensions: {sheet['maxRow']} rows x {sheet['maxColumn']} columns")
        print(f"Number of non-empty cells: {len(sheet['cells'])}")
        
        # Count different types of cells
        formulas = sum(1 for cell in sheet["cells"].values() if "formula" in cell)
        array_formulas = sum(1 for cell in sheet["cells"].values() if "array_formula" in cell)
        styled_cells = sum(1 for cell in sheet["cells"].values() 
                         if any(cell["style"]["font"].values()) or 
                            cell["style"]["fill"]["background"] is not None)
        
        print(f"Formulas: {formulas}")
        print(f"Array formulas: {array_formulas}")
        print(f"Styled cells: {styled_cells}")

if __name__ == "__main__":
    pytest.main(["-v", __file__]) 