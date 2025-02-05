#!/usr/bin/env python3

from openpyxl import load_workbook
from typing import Dict, Any, Optional, Set, List, Tuple
from termcolor import colored
import os
import sys
import re
import json
from datetime import datetime
from openpyxl.styles.colors import RGB

# Constants
OUTPUT_DIR = "output"
CELL_REF_PATTERN = re.compile(r'([A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?)')
NAMED_RANGE_PATTERN = re.compile(r'[A-Za-z][A-Za-z0-9_.]*')

class ExcelJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle Excel-specific types."""
    def default(self, obj):
        if isinstance(obj, RGB):
            # Convert RGB object to hex string or return None if no color
            if hasattr(obj, 'rgb'):
                return obj.rgb
            elif hasattr(obj, 'index'):
                # Handle theme colors
                return f"theme_{obj.index}"
            else:
                return None
        elif isinstance(obj, datetime):
            # Convert datetime objects to ISO format string
            return obj.isoformat()
        return super().default(obj)

def print_status(message: str, status: str = 'info') -> None:
    """Print colored status messages."""
    color_map = {
        'info': 'cyan',
        'success': 'green',
        'error': 'red',
        'warning': 'yellow'
    }
    print(colored(message, color_map.get(status, 'white')))

def ensure_output_dir() -> None:
    """Ensure output directory exists."""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print_status(f"Created output directory: {OUTPUT_DIR}", 'info')

def count_json_tokens(data: Dict[str, Any]) -> int:
    """
    Count the number of tokens in JSON data.
    This is an approximation that tries to match typical LLM tokenization patterns.
    """
    try:
        # Convert to JSON string
        json_str = json.dumps(data)
        
        # Basic tokenization rules that approximate LLM tokenizers:
        # 1. Split on whitespace
        # 2. Split on punctuation
        # 3. Split on case changes (camelCase, PascalCase)
        # 4. Keep common JSON structural elements as single tokens
        
        # First, handle JSON structural elements
        json_str = re.sub(r'([\{\}\[\],:])', r' \1 ', json_str)
        
        # Split on camelCase and PascalCase
        json_str = re.sub(r'([a-z])([A-Z])', r'\1 \2', json_str)
        json_str = re.sub(r'([A-Z])([A-Z][a-z])', r'\1 \2', json_str)
        
        # Handle numbers and special characters
        json_str = re.sub(r'([0-9]+)', r' \1 ', json_str)
        json_str = re.sub(r'([^a-zA-Z0-9\s\{\}\[\],:])', r' \1 ', json_str)
        
        # Split and filter empty strings
        tokens = [token for token in json_str.split() if token.strip()]
        
        # Add a small overhead for JSON structure and formatting
        structural_overhead = len(re.findall(r'[\{\}\[\],:]', json_str)) * 0.5
        
        # Return total token count with overhead
        return int(len(tokens) + structural_overhead)
        
    except Exception as e:
        print_status(f"Error counting tokens: {str(e)}", 'error')
        return 0

def get_output_filepath(excel_path: str, token_count: int) -> str:
    """Generate timestamped output filepath with token count."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    return os.path.join(OUTPUT_DIR, f"{base_name}_{token_count}tokens_{timestamp}.json")

def save_json_output(data: Dict[str, Any], output_path: str) -> None:
    """Save JSON data to file with proper formatting."""
    try:
        # Add metadata including token count
        if 'metadata' not in data:
            data['metadata'] = {}
        
        token_count = count_json_tokens(data)
        print_status(f"Total tokens in output: {token_count}", 'info')
        
        data['metadata'].update({
            'token_count': token_count,
            'conversion_timestamp': datetime.now().isoformat(),
            'original_filename': os.path.basename(output_path)
        })
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, cls=ExcelJSONEncoder, ensure_ascii=False)
            
        print_status(f"Successfully saved JSON to: {output_path} ({token_count} tokens)", 'success')
        
        # Print LLM recommendations
        print_status("\nLLM Recommendations based on token count:", 'info')
        if token_count < 4000:
            print_status("✓ Suitable for GPT-3.5-turbo (4K context)", 'success')
        elif token_count < 8000:
            print_status("✓ Suitable for GPT-3.5-turbo-8K", 'success')
        elif token_count < 16000:
            print_status("✓ Suitable for GPT-3.5-turbo-16K", 'success')
        elif token_count < 32000:
            print_status("✓ Suitable for GPT-4 (32K context)", 'success')
        else:
            print_status("⚠ Warning: Large token count. Consider splitting the data or using a model with larger context window", 'warning')
            
    except Exception as e:
        print_status(f"Error saving JSON output: {str(e)}", 'error')
        raise

def get_column_letter(col_num: int) -> str:
    """Convert column number to letter (1 = A, 2 = B, etc.)."""
    result = ""
    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def extract_cell_dependencies(formula: str) -> Set[str]:
    """Extract cell references from a formula."""
    if not formula:
        return set()
    
    try:
        # Find all cell references in the formula
        cell_refs = set(CELL_REF_PATTERN.findall(formula))
        # Remove any string literals that might look like cell references
        return {ref for ref in cell_refs if not ref.startswith('"') and not ref.startswith("'")}
    except Exception as e:
        print_status(f"Warning: Could not extract dependencies from formula: {str(e)}", 'warning')
        return set()

def process_array_formula(cell: Any) -> Dict[str, Any]:
    """Process array formula and its evaluation context."""
    try:
        if isinstance(cell.value, str) and cell.value.startswith('{'):
            return {
                "type": "array_formula",
                "formula": cell.value[1:-1],  # Remove { }
                "range": cell.coordinate,
                "calculated_value": cell.internal_value,
                "dimensions": {
                    "rows": len(cell.value.split(';')),
                    "columns": len(cell.value.split(';')[0].split(','))
                }
            }
        return None
    except Exception as e:
        print_status(f"Warning: Failed to process array formula: {str(e)}", 'warning')
        return None

def excel_to_json(file_path: str, sample_size: Optional[int] = None) -> Optional[Dict[str, Any]]:
    """
    Convert Excel file to JSON structure with optional row sampling.
    
    Args:
        file_path: Path to the Excel file
        sample_size: Optional maximum number of rows to process per sheet
        
    Returns:
        Dict containing the spreadsheet structure or None if an error occurs
    """
    try:
        # Validate file existence
        if not os.path.exists(file_path):
            print_status(f"Error: File not found: {file_path}", 'error')
            return None

        # Validate file extension
        if not any(file_path.lower().endswith(ext) for ext in ['.xlsx', '.xls', '.xlsm']):
            print_status(f"Error: Unsupported file format. File must be .xlsx, .xls, or .xlsm", 'error')
            return None

        print_status(f"Loading workbook: {file_path}", 'info')
        try:
            wb = load_workbook(filename=file_path, data_only=False)
        except Exception as e:
            print_status(f"Error: Failed to load Excel file: {str(e)}", 'error')
            return None
        
        spreadsheet_dict = {
            "fileName": os.path.basename(file_path),
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            try:
                print_status(f"Processing sheet: {sheet_name}", 'info')
                sheet = wb[sheet_name]
                sheet_data = {
                    "sheetTitle": sheet_name,
                    "maxRow": sheet.max_row,
                    "maxColumn": sheet.max_column,
                    "cells": {}
                }

                # Determine how many rows we should actually iterate
                row_limit = sheet.max_row if sample_size is None else min(sample_size, sheet.max_row)
                
                if sample_size:
                    print_status(f"Processing {row_limit} rows out of {sheet.max_row} total rows", 'info')

                # Process cells
                for row_index, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    if row_index > row_limit:
                        break

                    for cell in row:
                        try:
                            cell_coord = cell.coordinate
                            cell_data = {"value": cell.value}

                            # Handle array formulas
                            array_formula = process_array_formula(cell)
                            if array_formula:
                                cell_data["array_formula"] = array_formula

                            # Handle regular formulas
                            if cell.data_type == "f":
                                cell_data.update({
                                    "formula": cell.value,
                                    "calculated_value": cell.internal_value,
                                    "dependencies": list(extract_cell_dependencies(cell.value))
                                })

                            # Add style information
                            cell_data.update({
                                "style": {
                                    "font": {
                                        "bold": cell.font.bold,
                                        "italic": cell.font.italic,
                                        "color": cell.font.color.rgb if cell.font.color else None
                                    },
                                    "fill": {
                                        "background": cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill, 'start_color') else None
                                    },
                                    "alignment": {
                                        "horizontal": cell.alignment.horizontal,
                                        "vertical": cell.alignment.vertical
                                    }
                                }
                            })

                            # Only add non-empty cells to reduce output size
                            if cell.value is not None:
                                sheet_data["cells"][cell_coord] = cell_data

                        except Exception as e:
                            print_status(f"Warning: Failed to process cell {cell_coord} in sheet {sheet_name}: {str(e)}", 'warning')
                            continue

                spreadsheet_dict["sheets"].append(sheet_data)
                print_status(f"Successfully processed sheet: {sheet_name}", 'success')

            except Exception as e:
                print_status(f"Warning: Failed to process sheet {sheet_name}: {str(e)}", 'warning')
                continue

        if not spreadsheet_dict["sheets"]:
            print_status("Error: No valid sheets were processed", 'error')
            return None

        # Count tokens and get output path
        token_count = count_json_tokens(spreadsheet_dict)
        output_path = get_output_filepath(file_path, token_count)
        
        # Save with token count in filename
        save_json_output(spreadsheet_dict, output_path)

        print_status("Successfully converted Excel file to JSON structure", 'success')
        return spreadsheet_dict

    except Exception as e:
        print_status(f"Error converting Excel file: {str(e)}", 'error')
        return None

if __name__ == "__main__":
    # Example usage
    if len(sys.argv) > 1:
        try:
            result = excel_to_json(sys.argv[1], sample_size=int(sys.argv[2]) if len(sys.argv) > 2 else None)
            if result:
                sys.exit(0)
            else:
                sys.exit(1)
        except Exception as e:
            print_status(f"Error: {str(e)}", 'error')
            sys.exit(1)
    else:
        print_status("Usage: python excel_converter.py <excel_file> [sample_size]", 'info')
        sys.exit(1) 