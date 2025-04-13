#!/usr/bin/env python3

import json
import os
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from termcolor import colored
import sys
import datetime
import re
import collections

# Constants
SUPPORTED_EXTENSIONS = ['.xlsx', '.xls', '.xlsm']
OUTPUT_DIR = 'converted_json'
DEFAULT_ENCODING = 'utf-8'

# Token-efficient property names
TOKEN_EFFICIENT = True  # Set to True to use abbreviated property names
PROPERTY_MAP = {
    'value': 'v',
    'formula': 'f',
    'calculated_value': 'cv',
    'style': 's',
    'data_validation': 'dv',
    'hyperlink': 'h',
    'comment': 'c',
    'dependencies': 'd',
    'metadata': 'm',
    'cells': 'cl',
    'merged_cells': 'mc',
    'conditional_formatting': 'cf',
    'protection': 'p',
    'view_settings': 'vs',
    'pivot_tables': 'pt',
    'charts': 'ch',
    'form_controls': 'fc',
    'auto_filter': 'af',
    'sheets': 'sh',
    'file_name': 'fn',
    'named_ranges': 'nr',
    'title': 't',
    'dimensions': 'dim',
    'max_row': 'mr',
    'max_column': 'mc',
    'sheet_properties': 'sp',
    'token_count': 'tc',
    'conversion_timestamp': 'ts',
    'original_filename': 'of',
    'tables': 'tb',
    'samples': 'sp',
    'column_types': 'ct',
    'formula_patterns': 'fp',
    'implementation_notes': 'in',
    'range': 'rg',
    'columns': 'cols',
    'header_row': 'hr',
    'js_equivalent': 'js',
    'structured_references': 'sr',
    'enriched_context': 'ec'
}

class ExcelJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle Excel-specific types."""
    def default(self, obj):
        try:
            if isinstance(obj, datetime.datetime):
                return obj.isoformat()
            elif isinstance(obj, datetime.date):
                return obj.isoformat()
            elif hasattr(obj, 'rgb'):
                # Handle RGB objects from openpyxl
                if obj.rgb:
                    return obj.rgb.decode() if isinstance(obj.rgb, bytes) else str(obj.rgb)
                return None
            elif hasattr(obj, '__dict__'):
                # Handle other objects by converting their attributes to dict
                return {k: v for k, v in obj.__dict__.items() 
                       if not k.startswith('_') and v is not None}
            return super().default(obj)
        except Exception as e:
            # If we can't serialize, return None instead of failing
            print_status(f"Warning: Could not serialize object {type(obj)}: {str(e)}", 'warning')
            return None

def print_status(message: str, status: str = 'info') -> None:
    """Print colored status messages."""
    color_map = {
        'info': 'cyan',
        'success': 'green',
        'error': 'red',
        'warning': 'yellow'
    }
    print(colored(message, color_map.get(status, 'white')))

def map_key(key: str) -> str:
    """Map a key to its token-efficient version if TOKEN_EFFICIENT is True."""
    if TOKEN_EFFICIENT and key in PROPERTY_MAP:
        return PROPERTY_MAP[key]
    return key

def get_data_validation(cell: Any) -> Optional[Dict[str, Any]]:
    """Extract data validation information from a cell."""
    try:
        if hasattr(cell, 'data_validation') and cell.data_validation:
            validation_data = {
                'type': cell.data_validation.type,
                'operator': cell.data_validation.operator,
                'formula1': cell.data_validation.formula1,
                'formula2': cell.data_validation.formula2,
                'allow_blank': cell.data_validation.allow_blank,
                'show_error_message': cell.data_validation.showErrorMessage,
                'error_title': cell.data_validation.errorTitle,
                'error_message': cell.data_validation.error,
                'prompt_title': cell.data_validation.promptTitle,
                'prompt_message': cell.data_validation.prompt,
            }
            return validation_data
        return None
    except Exception as e:
        # Don't print error for missing data validation
        return None

def get_cell_value(cell: Any) -> Any:
    """Extract cell value and handle different data types."""
    try:
        if cell.value is None:
            return None
        
        # Handle formula
        if cell.data_type == 'f':
            if TOKEN_EFFICIENT:
                return {
                    map_key('formula'): str(cell.value),
                    map_key('calculated_value'): cell.internal_value
                }
            else:
                return {
                    'formula': str(cell.value),
                    'calculated_value': cell.internal_value
                }
        
        # Handle dates
        if isinstance(cell.value, (datetime.date, datetime.datetime)):
            return cell.value.isoformat()
            
        return cell.value
    except Exception as e:
        print_status(f"Error processing cell: {str(e)}", 'error')
        return None

def get_sheet_metadata(sheet: Worksheet) -> Dict[str, Any]:
    """Extract minimal sheet metadata."""
    try:
        if TOKEN_EFFICIENT:
            return {
                map_key('title'): sheet.title,
                map_key('dimensions'): sheet.dimensions,
                map_key('max_row'): sheet.max_row,
                map_key('max_column'): sheet.max_column
            }
        else:
            return {
                'title': sheet.title,
                'dimensions': sheet.dimensions,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column
            }
    except Exception as e:
        print_status(f"Error processing sheet metadata: {str(e)}", 'error')
        if TOKEN_EFFICIENT:
            return {
                map_key('title'): sheet.title,
                map_key('dimensions'): sheet.dimensions,
                map_key('max_row'): sheet.max_row,
                map_key('max_column'): sheet.max_column
            }
        else:
            return {
                'title': sheet.title,
                'dimensions': sheet.dimensions,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column
            }

def get_conditional_formatting(sheet: Worksheet) -> List[Dict[str, Any]]:
    """Extract conditional formatting rules from the sheet."""
    try:
        cf_rules = []
        for cf_range in sheet.conditional_formatting:
            for rule in cf_range.rules:
                cf_rules.append({
                    'range': str(cf_range),
                    'type': rule.type,
                    'priority': rule.priority,
                    'formula': rule.formula if hasattr(rule, 'formula') else None,
                    'dxf': {
                        'font': rule.dxf.font.__dict__ if rule.dxf and hasattr(rule.dxf, 'font') else None,
                        'fill': rule.dxf.fill.__dict__ if rule.dxf and hasattr(rule.dxf, 'fill') else None,
                    } if hasattr(rule, 'dxf') and rule.dxf else None
                })
        return cf_rules
    except Exception as e:
        print_status(f"Error processing conditional formatting: {str(e)}", 'error')
        return []

def get_named_ranges(workbook: Any) -> Dict[str, Any]:
    """Extract named ranges from the workbook."""
    try:
        named_ranges = {}
        if hasattr(workbook, 'defined_names'):
            # Handle different versions of openpyxl
            if hasattr(workbook.defined_names, 'items'):
                # New version of openpyxl
                for name, defn in workbook.defined_names.items():
                    try:
                        # Get the destinations for this named range
                        destinations = defn.destinations if hasattr(defn, 'destinations') else [(defn.attr_map.get('localSheetId'), defn.value)]
                        
                        # Process each destination
                        range_refs = []
                        for sheet_id, coord in destinations:
                            try:
                                # Get sheet name if sheet_id is provided
                                sheet_name = None
                                if sheet_id is not None:
                                    sheet_name = workbook.sheetnames[int(sheet_id)]
                                
                                # Clean up the reference
                                clean_ref = coord.strip('$')
                                if '!' not in clean_ref and sheet_name:
                                    clean_ref = f"{sheet_name}!{clean_ref}"
                                elif '!' not in clean_ref:
                                    # Use the first sheet if no sheet is specified
                                    clean_ref = f"{workbook.sheetnames[0]}!{clean_ref}"
                                
                                range_refs.append(clean_ref)
                            except Exception as e:
                                print_status(f"Warning: Could not process destination for named range '{name}': {str(e)}", 'warning')
                                continue
                        
                        if range_refs:
                            named_ranges[name] = {
                                'value': range_refs[0] if len(range_refs) == 1 else range_refs,
                                'scope': 'workbook' if defn.localSheetId is None else f"sheet_{defn.localSheetId}"
                            }
                    except Exception as e:
                        print_status(f"Warning: Could not process named range '{name}': {str(e)}", 'warning')
                        continue
            
            elif hasattr(workbook.defined_names, '_dict'):
                # Alternative version of openpyxl
                for name, defn in workbook.defined_names._dict.items():
                    try:
                        # Get the reference
                        ref = defn.value if hasattr(defn, 'value') else defn.attr_map.get('refersTo', '')
                        
                        # Clean up the reference
                        if ref.startswith('='):
                            ref = ref[1:]
                        ref = ref.strip('$')
                        
                        # Add sheet name if missing
                        if '!' not in ref:
                            sheet_id = defn.localSheetId if hasattr(defn, 'localSheetId') else None
                            sheet_name = workbook.sheetnames[int(sheet_id)] if sheet_id is not None else workbook.sheetnames[0]
                            ref = f"{sheet_name}!{ref}"
                        
                        named_ranges[name] = {
                            'value': ref,
                            'scope': 'workbook' if not hasattr(defn, 'localSheetId') or defn.localSheetId is None else f"sheet_{defn.localSheetId}"
                        }
                    except Exception as e:
                        print_status(f"Warning: Could not process named range '{name}': {str(e)}", 'warning')
                        continue
        
        return named_ranges
    except Exception as e:
        print_status(f"Warning: Could not process named ranges: {str(e)}", 'warning')
        return {}

def get_protection_settings(sheet: Worksheet) -> Dict[str, Any]:
    """Extract protection settings from the sheet."""
    try:
        protection_info = {}
        
        # Check if sheet has protection
        if hasattr(sheet, 'protection') and sheet.protection:
            protection = sheet.protection
            protection_info['sheet_protection'] = {
                'enabled': getattr(protection, 'sheet', False),
                'password': getattr(protection, 'password', None) is not None,
            }
            
            # Get available protection attributes
            for attr in dir(protection):
                if not attr.startswith('_') and attr not in ['sheet', 'password']:
                    protection_info['sheet_protection'][attr] = getattr(protection, attr, None)
        
        # Get protected ranges if available
        if hasattr(sheet, 'protected_ranges'):
            protection_info['protected_ranges'] = [
                str(protected_range) for protected_range in sheet.protected_ranges
            ]
        else:
            protection_info['protected_ranges'] = []
            
        return protection_info
    except Exception as e:
        print_status(f"Warning: Could not process protection settings: {str(e)}", 'warning')
        return {
            'sheet_protection': {'enabled': False},
            'protected_ranges': []
        }

def get_hyperlinks(cell: Any) -> Optional[Dict[str, Any]]:
    """Extract hyperlink information from a cell."""
    try:
        if cell.hyperlink:
            return {
                'target': cell.hyperlink.target,
                'tooltip': cell.hyperlink.tooltip if hasattr(cell.hyperlink, 'tooltip') else None,
                'location': cell.hyperlink.location if hasattr(cell.hyperlink, 'location') else None
            }
        return None
    except Exception as e:
        print_status(f"Error processing hyperlink: {str(e)}", 'error')
        return None

def get_comments(cell: Any) -> Optional[Dict[str, Any]]:
    """Extract comment information from a cell."""
    try:
        if cell.comment:
            return {
                'text': cell.comment.text,
                'author': cell.comment.author if hasattr(cell.comment, 'author') else None
            }
        return None
    except Exception as e:
        print_status(f"Error processing comment: {str(e)}", 'error')
        return None

def get_sheet_view_settings(sheet: Worksheet) -> Dict[str, Any]:
    """Extract sheet view settings."""
    try:
        return {
            'frozen_panes': {
                'rows': sheet.freeze_panes[1] if sheet.freeze_panes else None,
                'columns': sheet.freeze_panes[0] if sheet.freeze_panes else None
            },
            'zoom_scale': sheet.sheet_view.zoomScale if hasattr(sheet.sheet_view, 'zoomScale') else 100,
            'show_gridlines': sheet.sheet_view.showGridLines if hasattr(sheet.sheet_view, 'showGridLines') else True,
            'hidden_rows': [row for row in range(1, sheet.max_row + 1) if sheet.row_dimensions[row].hidden],
            'hidden_columns': [col for col in range(1, sheet.max_column + 1) if sheet.column_dimensions[get_column_letter(col)].hidden]
        }
    except Exception as e:
        print_status(f"Error processing sheet view settings: {str(e)}", 'error')
        return {}

def get_pivot_tables(sheet: Worksheet) -> List[Dict[str, Any]]:
    """Extract pivot table information from the sheet in a token-efficient way."""
    try:
        pivot_tables = []
        if hasattr(sheet, '_pivots'):
            for pivot in sheet._pivots:
                # Create a token-efficient pivot table representation
                if TOKEN_EFFICIENT:
                    pivot_data = {
                        'loc': str(pivot.location) if hasattr(pivot, 'location') else None,
                        'nm': pivot.name if hasattr(pivot, 'name') else None
                    }
                    
                    # Extract fields using token-efficient keys
                    fields = {}
                    
                    # Row fields
                    if hasattr(pivot, 'row_fields') and pivot.row_fields:
                        fields['rf'] = [{'nm': field.name} for field in pivot.row_fields if hasattr(field, 'name')]
                    
                    # Column fields
                    if hasattr(pivot, 'column_fields') and pivot.column_fields:
                        fields['cf'] = [{'nm': field.name} for field in pivot.column_fields if hasattr(field, 'name')]
                    
                    # Data fields with aggregation functions
                    if hasattr(pivot, 'data_fields') and pivot.data_fields:
                        fields['df'] = [
                            {
                                'nm': field.name if hasattr(field, 'name') else None,
                                'fn': field.function if hasattr(field, 'function') else None
                            } 
                            for field in pivot.data_fields
                        ]
                    
                    # Page/Filter fields
                    if hasattr(pivot, 'page_fields') and pivot.page_fields:
                        fields['pf'] = [{'nm': field.name} for field in pivot.page_fields if hasattr(field, 'name')]
                    
                    # Add fields if they exist
                    if fields:
                        pivot_data['fld'] = fields
                        
                    # Add display options if relevant
                    options = {}
                    if hasattr(pivot, 'merge_labels') and pivot.merge_labels is not None:
                        options['ml'] = pivot.merge_labels
                    if hasattr(pivot, 'show_error') and pivot.show_error is not None:
                        options['se'] = pivot.show_error
                    if hasattr(pivot, 'show_empty') and pivot.show_empty is not None:
                        options['em'] = pivot.show_empty
                        
                    # Add options if they exist
                    if options:
                        pivot_data['opt'] = options
                        
                else:
                    # Non-token-efficient representation
                    pivot_data = {
                        'location': str(pivot.location) if hasattr(pivot, 'location') else None,
                        'name': pivot.name if hasattr(pivot, 'name') else None
                    }
                    
                    # Extract fields
                    fields = {}
                    
                    # Row fields
                    if hasattr(pivot, 'row_fields') and pivot.row_fields:
                        fields['row_fields'] = [{'name': field.name} for field in pivot.row_fields if hasattr(field, 'name')]
                    
                    # Column fields
                    if hasattr(pivot, 'column_fields') and pivot.column_fields:
                        fields['column_fields'] = [{'name': field.name} for field in pivot.column_fields if hasattr(field, 'name')]
                    
                    # Data fields with aggregation functions
                    if hasattr(pivot, 'data_fields') and pivot.data_fields:
                        fields['data_fields'] = [
                            {
                                'name': field.name if hasattr(field, 'name') else None,
                                'function': field.function if hasattr(field, 'function') else None
                            } 
                            for field in pivot.data_fields
                        ]
                    
                    # Page/Filter fields
                    if hasattr(pivot, 'page_fields') and pivot.page_fields:
                        fields['page_fields'] = [{'name': field.name} for field in pivot.page_fields if hasattr(field, 'name')]
                    
                    # Add fields if they exist
                    if fields:
                        pivot_data['fields'] = fields
                        
                    # Add display options if relevant
                    options = {}
                    if hasattr(pivot, 'merge_labels') and pivot.merge_labels is not None:
                        options['merge_labels'] = pivot.merge_labels
                    if hasattr(pivot, 'show_error') and pivot.show_error is not None:
                        options['show_error'] = pivot.show_error
                    if hasattr(pivot, 'show_empty') and pivot.show_empty is not None:
                        options['show_empty'] = pivot.show_empty
                        
                    # Add options if they exist
                    if options:
                        pivot_data['options'] = options
                        
                pivot_tables.append(pivot_data)
                
                # Limit to 5 pivot tables per sheet for token efficiency
                if len(pivot_tables) >= 5:
                    break
        
        # Try alternative approach for newer Excel/openpyxl versions
        if not pivot_tables and hasattr(sheet, 'pivotTables') and sheet.pivotTables:
            for pivot in sheet.pivotTables:
                if TOKEN_EFFICIENT:
                    pivot_data = {
                        'loc': str(pivot.location) if hasattr(pivot, 'location') else None,
                        'nm': pivot.name if hasattr(pivot, 'name') else None
                    }
                else:
                    pivot_data = {
                        'location': str(pivot.location) if hasattr(pivot, 'location') else None,
                        'name': pivot.name if hasattr(pivot, 'name') else None
                    }
                
                pivot_tables.append(pivot_data)
                
                # Limit to 5 pivot tables per sheet for token efficiency
                if len(pivot_tables) >= 5:
                    break
                
        return pivot_tables
    except Exception as e:
        print_status(f"Warning: Could not process pivot tables: {str(e)}", 'warning')
        return []

def get_charts(sheet: Worksheet) -> List[Dict[str, Any]]:
    """Extract chart information from the sheet."""
    try:
        charts = []
        if hasattr(sheet, '_charts'):
            for chart in sheet._charts:
                chart_data = {
                    'type': type(chart).__name__,
                    'title': chart.title.text if hasattr(chart, 'title') and chart.title else None,
                    'anchor': str(chart.anchor) if hasattr(chart, 'anchor') else None,
                    'series': [{
                        'title': series.title if hasattr(series, 'title') else None,
                        'values': str(series.values) if hasattr(series, 'values') else None,
                        'categories': str(series.categories) if hasattr(series, 'categories') else None
                    } for series in chart.series] if hasattr(chart, 'series') else [],
                    'style': chart.style if hasattr(chart, 'style') else None,
                    'legend': {
                        'position': chart.legend.position if hasattr(chart, 'legend') and hasattr(chart.legend, 'position') else None,
                        'overlay': chart.legend.overlay if hasattr(chart, 'legend') and hasattr(chart.legend, 'overlay') else None
                    } if hasattr(chart, 'legend') else None
                }
                charts.append(chart_data)
        return charts
    except Exception as e:
        print_status(f"Error processing charts: {str(e)}", 'error')
        return []

def get_form_controls(sheet: Worksheet) -> List[Dict[str, Any]]:
    """Extract form control information from the sheet."""
    try:
        controls = []
        if hasattr(sheet, '_controls'):
            for control in sheet._controls:
                control_data = {
                    'type': type(control).__name__,
                    'name': control.name if hasattr(control, 'name') else None,
                    'location': str(control.anchor) if hasattr(control, 'anchor') else None,
                    'properties': {
                        'caption': control.caption if hasattr(control, 'caption') else None,
                        'value': control.value if hasattr(control, 'value') else None,
                        'linked_cell': str(control.linked_cell) if hasattr(control, 'linked_cell') else None,
                        'disabled': control.disabled if hasattr(control, 'disabled') else None,
                        'print_object': control.print_object if hasattr(control, 'print_object') else None,
                        'macro': control.macro if hasattr(control, 'macro') else None
                    }
                }
                controls.append(control_data)
        return controls
    except Exception as e:
        print_status(f"Error processing form controls: {str(e)}", 'error')
        return []

def get_auto_filters(sheet: Worksheet) -> Optional[Dict[str, Any]]:
    """Extract auto-filter information from the sheet."""
    try:
        if sheet.auto_filter.ref:
            filter_data = {
                'range': str(sheet.auto_filter.ref),
                'filters': {}
            }
            
            if hasattr(sheet.auto_filter, 'filterColumn'):
                for col_id, filter_column in sheet.auto_filter.filterColumn.items():
                    filter_data['filters'][col_id] = {
                        'type': filter_column.type if hasattr(filter_column, 'type') else None,
                        'values': filter_column.vals if hasattr(filter_column, 'vals') else None,
                        'custom_filters': filter_column.customFilters if hasattr(filter_column, 'customFilters') else None
                    }
            return filter_data
        return None
    except Exception as e:
        print_status(f"Error processing auto-filters: {str(e)}", 'error')
        return None

def get_cell_dependencies(cell: Any, sheet: Worksheet) -> Optional[Dict[str, Any]]:
    """Extract formula dependencies from a cell."""
    try:
        if cell.data_type == 'f' and cell.value:
            formula = str(cell.value)
            
            # Find all cell references in the formula (e.g., A1, B2:C3)
            cell_refs = re.findall(r'[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?', formula)
            
            # Find all named ranges in the formula
            named_ranges = re.findall(r'[a-zA-Z][a-zA-Z0-9_.]*(?=\W)', formula)
            
            # Remove common Excel functions from named ranges
            common_functions = {'SUM', 'IF', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'AND', 'OR', 'NOT', 'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'OFFSET', 'INDIRECT', 'ROW', 'COLUMN', 'CELL', 'IFERROR', 'ISERROR', 'ISBLANK', 'ISTEXT', 'ISNONTEXT', 'ISNUMBER', 'TODAY', 'NOW', 'TEXT', 'VALUE', 'LEN', 'LEFT', 'RIGHT', 'MID', 'TRIM', 'CONCATENATE', 'CONCAT', 'SUBSTITUTE', 'REPLACE', 'DATE', 'DAY', 'MONTH', 'YEAR', 'EOMONTH', 'NETWORKDAYS', 'WORKDAY', 'TRUE', 'FALSE'}
            named_ranges = [nr for nr in named_ranges if nr not in common_functions]
            
            if TOKEN_EFFICIENT:
                if cell_refs or named_ranges:
                    result = {}
                    if cell_refs:
                        result['cr'] = cell_refs
                    if named_ranges:
                        result['nr'] = named_ranges
                    return result
                return None
            else:
                if cell_refs or named_ranges:
                    result = {}
                    if cell_refs:
                        result['cell_references'] = cell_refs
                    if named_ranges:
                        result['named_ranges'] = named_ranges
                    return result
                return None
        return None
    except Exception as e:
        # Don't print error for missing dependencies
        return None

def extract_table_structures(workbook):
    """Extract Excel table structures (ListObjects)."""
    tables = {}
    
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Check if sheet has tables defined
            if hasattr(sheet, '_tables') and sheet._tables:
                for table in sheet._tables:
                    table_name = table.name if hasattr(table, 'name') else f"Table_{len(tables)+1}"
                    
                    # Get table range
                    table_range = str(table.ref) if hasattr(table, 'ref') else None
                    
                    # Get table columns with correct column letters
                    columns = []
                    if hasattr(table, 'tableColumns') and table_range:
                        # Parse the range to get starting column
                        range_match = re.match(r'([A-Z]+)(\d+):([A-Z]+)', table_range)
                        if range_match:
                            start_col_letter = range_match.group(1)
                            start_col_index = openpyxl.utils.column_index_from_string(start_col_letter)
                            
                            for i, col in enumerate(table.tableColumns):
                                if hasattr(col, 'name'):
                                    # Use actual Excel column letters
                                    col_letter = get_column_letter(start_col_index + i)
                                    columns.append({
                                        'name': col.name,
                                        'col': col_letter
                                    })
                    
                    # Infer header row from range if possible
                    header_row = None
                    if table_range:
                        match = re.match(r'[A-Z]+(\d+):', table_range)
                        if match:
                            header_row = int(match.group(1))
                    
                    if TOKEN_EFFICIENT:
                        tables[table_name] = {
                            map_key('range'): table_range,
                            map_key('columns'): columns,
                            map_key('header_row'): header_row
                        }
                    else:
                        tables[table_name] = {
                            'range': table_range,
                            'columns': columns,
                            'header_row': header_row
                        }
            
            # Collect structured references from all formulas across the workbook
            structured_refs = {}
            
            # First pass: collect all table and column references from formulas
            for sheet_scan_name in workbook.sheetnames:
                sheet_scan = workbook[sheet_scan_name]
                
                for row in sheet_scan.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f' and cell.value:
                            # Find structured references like TableName[[#This Row],[ColumnName]]
                            matches = re.findall(r'([A-Za-z0-9_]+)\[\[#.*?\],\[([^\]]+)\]\]', str(cell.value))
                            for table_name, column_name in matches:
                                if table_name not in structured_refs:
                                    structured_refs[table_name] = set()
                                structured_refs[table_name].add(column_name)
            
            # Second pass: process each table reference and try to locate the actual columns
            for table_name, columns_referenced in structured_refs.items():
                if table_name not in tables:  # Don't override real tables
                    # Find header row containing these column names
                    header_row = None
                    column_positions = {}  # Map column names to their positions
                    
                    # Scan for a row with matching column headers
                    for row_idx in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
                        matches_found = 0
                        match_positions = {}
                        
                        for col_idx in range(1, min(sheet.max_column + 1, 30)):  # Check first 30 columns
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            if cell.value in columns_referenced:
                                matches_found += 1
                                match_positions[cell.value] = col_idx
                        
                        # If we found at least one match, assume this is the header row
                        if matches_found > 0 and matches_found >= len(columns_referenced) * 0.5:  # At least half the columns
                            header_row = row_idx
                            column_positions = match_positions
                            break
                    
                    # Build columns list with proper column letters
                    columns = []
                    if column_positions:
                        for col_name, col_idx in column_positions.items():
                            col_letter = get_column_letter(col_idx)
                            columns.append({
                                'name': col_name,
                                'col': col_letter
                            })
                    else:
                        # Fallback: just create entries with proper names, but no columns
                        for col_name in columns_referenced:
                            columns.append({
                                'name': col_name,
                                'col': None
                            })
                    
                    # Determine table range - include all columns that are part of the table
                    table_range = None
                    if column_positions and header_row:
                        col_indices = list(column_positions.values())
                        if col_indices:
                            min_col = min(col_indices)
                            max_col = max(col_indices)
                            
                            # Estimate table height - assume at least 100 rows or to the end of sheet data
                            max_data_row = header_row + 100
                            for col_idx in range(min_col, max_col + 1):
                                col_letter = get_column_letter(col_idx)
                                # Check for data in this column
                                for row_idx in range(header_row + 1, sheet.max_row + 1):
                                    if sheet.cell(row=row_idx, column=col_idx).value is not None:
                                        max_data_row = max(max_data_row, row_idx)
                            
                            start_col_letter = get_column_letter(min_col)
                            end_col_letter = get_column_letter(max_col)
                            table_range = f"{start_col_letter}{header_row}:{end_col_letter}{max_data_row}"
                    
                    if TOKEN_EFFICIENT:
                        tables[table_name] = {
                            map_key('range'): table_range or "Inferred",
                            map_key('columns'): columns,
                            map_key('header_row'): header_row
                        }
                    else:
                        tables[table_name] = {
                            'range': table_range or "Inferred",
                            'columns': columns,
                            'header_row': header_row
                        }
    
    except Exception as e:
        print_status(f"Warning: Could not extract table structures: {str(e)}", 'warning')
    
    # Remove empty or unused tables
    tables_to_remove = []
    for table_name, table_info in tables.items():
        cols_key = map_key('columns') if TOKEN_EFFICIENT else 'columns'
        if not table_info[cols_key]:
            tables_to_remove.append(table_name)
    
    for table_name in tables_to_remove:
        del tables[table_name]
    
    return tables

def extract_formula_patterns(data):
    """Identify repeating formula patterns in the workbook."""
    formula_patterns = {}
    formula_counts = collections.Counter()
    
    try:
        # Count formula occurrences
        for sheet_name, sheet_data in data[map_key('sheets') if TOKEN_EFFICIENT else 'sheets'].items():
            for cell_ref, cell_data in sheet_data[map_key('cells') if TOKEN_EFFICIENT else 'cells'].items():
                cell_value = cell_data[map_key('value') if TOKEN_EFFICIENT else 'value']
                
                if isinstance(cell_value, dict) and (map_key('formula') if TOKEN_EFFICIENT else 'formula') in cell_value:
                    formula = cell_value[map_key('formula') if TOKEN_EFFICIENT else 'formula']
                    formula_counts[formula] += 1
        
        # Extract patterns (formulas used multiple times)
        for formula, count in formula_counts.items():
            if count > 1:
                pattern_name = f"pattern_{len(formula_patterns)+1}"
                formula_patterns[pattern_name] = formula
                
                # Limit to top 10 patterns to save tokens
                if len(formula_patterns) >= 10:
                    break
    
    except Exception as e:
        print_status(f"Warning: Could not extract formula patterns: {str(e)}", 'warning')
    
    return formula_patterns

def extract_column_types(workbook):
    """Extract column data types based on cell formatting."""
    column_types = {}
    
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Check a sample of cells in each column to infer type
            for col in range(1, min(sheet.max_column + 1, 20)):  # Limit to first 20 columns
                col_letter = get_column_letter(col)
                
                # Check header row (usually row 1 or 2)
                for header_row in [1, 2]:
                    header_cell = sheet.cell(row=header_row, column=col)
                    if header_cell.value:
                        # Skip if this is just a header
                        break
                
                # Sample data from a few rows
                data_type = None
                format_type = None
                
                for row in range(header_row + 1, min(sheet.max_row + 1, header_row + 5)):
                    cell = sheet.cell(row=row, column=col)
                    
                    # Check value type
                    if isinstance(cell.value, (int, float)):
                        data_type = 'number'
                        
                        # Check for currency format
                        if cell.number_format and ('$' in cell.number_format or '€' in cell.number_format):
                            format_type = 'currency'
                        elif cell.number_format and ('%' in cell.number_format):
                            format_type = 'percentage'
                        else:
                            format_type = 'decimal'
                    
                    elif isinstance(cell.value, datetime.datetime):
                        data_type = 'datetime'
                    elif isinstance(cell.value, datetime.date):
                        data_type = 'date'
                    elif isinstance(cell.value, str):
                        data_type = 'string'
                    
                    if data_type:
                        break
                
                if data_type:
                    column_types[col_letter] = format_type if format_type else data_type
                
                # Limit to save tokens
                if len(column_types) >= 15:  # Only store types for 15 columns
                    break
    
    except Exception as e:
        print_status(f"Warning: Could not extract column types: {str(e)}", 'warning')
    
    return column_types

def extract_sample_calculated_values(workbook, formula_cells, max_samples=10):
    """Extract a small set of calculated values for validation."""
    samples = {}
    
    try:
        # Create a data_only version of the workbook to get calculated values
        # This requires re-loading the workbook with data_only=True
        if hasattr(workbook, '_path'):
            data_workbook = openpyxl.load_workbook(workbook._path, data_only=True)
            
            # Limit to a small number of samples
            sample_count = 0
            for sheet_name, cells in formula_cells.items():
                if sheet_name in data_workbook.sheetnames:
                    data_sheet = data_workbook[sheet_name]
                    
                    for cell_ref in cells:
                        if sample_count >= max_samples:
                            break
                            
                        # Parse cell reference (e.g. "A1" to row 1, column 1)
                        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
                        if match:
                            col_letter, row_num = match.groups()
                            row = int(row_num)
                            
                            try:
                                data_cell = data_sheet[cell_ref]
                                if data_cell.value is not None:
                                    samples[cell_ref] = data_cell.value
                                    sample_count += 1
                            except:
                                continue
                
                if sample_count >= max_samples:
                    break
    except Exception as e:
        print_status(f"Warning: Could not extract calculated values: {str(e)}", 'warning')
    
    return samples

def get_implementation_notes(formula_patterns=None):
    """Generate implementation notes for the JavaScript developer."""
    
    # Create a dictionary of implementation notes
    notes = {}
    
    # Add structured reference explanation
    if TOKEN_EFFICIENT:
        notes[map_key('structured_references')] = "Table[[#This Row],[Column]] refers to the value in the current row of the specified column in the Excel table"
    else:
        notes['structured_references'] = "Table[[#This Row],[Column]] refers to the value in the current row of the specified column in the Excel table"
    
    # Add JavaScript implementation example
    if formula_patterns and len(formula_patterns) > 0:
        # Get the first formula pattern
        pattern_key = next(iter(formula_patterns))
        pattern = formula_patterns[pattern_key]
        
        # Try to create a reasonable JavaScript implementation
        js_implementation = "row => "
        
        # Replace common patterns in formulas with JavaScript equivalents
        if "[[#This Row],[" in pattern:
            # This is a structured reference to a table
            # Extract table and column references
            table_refs = re.findall(r'([A-Za-z0-9_]+)\[\[#This Row\],\[([^\]]+)\]\]', pattern)
            
            if table_refs:
                # Create a JavaScript expression using row object
                js_expression = pattern
                for table_name, column_name in table_refs:
                    js_expression = js_expression.replace(
                        f"{table_name}[[#This Row],[{column_name}]]", 
                        f"row['{column_name}']"
                    )
                
                # Replace basic Excel operators with JavaScript equivalents
                js_expression = js_expression.replace("=", "")
                js_expression = js_expression.replace("<>", "!==")
                js_expression = js_expression.replace("&", "+")  # String concatenation
                
                js_implementation += js_expression
            else:
                js_implementation += "/* Formula could not be automatically converted */"
        else:
            # Regular cell references - try a simple conversion
            js_implementation += "/* Formula requires custom implementation */"
            
        if TOKEN_EFFICIENT:
            notes[map_key('js_equivalent')] = js_implementation
        else:
            notes['js_equivalent'] = js_implementation
    
    # Add pivot table usage note if needed
    if TOKEN_EFFICIENT:
        notes[map_key('pivot_tables')] = "Pivot tables provide summarized views of data. In a web implementation, these can be created dynamically from the raw data."
    else:
        notes['pivot_tables'] = "Pivot tables provide summarized views of data. In a web implementation, these can be created dynamically from the raw data."
    
    return notes

def extract_data_validation_rules(workbook):
    """Extract data validation rules efficiently."""
    validation_rules = {}
    
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_rules = []
            
            # Track already processed rules to avoid duplicates
            processed_rules = set()
            
            # Sample rows and columns for validation rules
            max_rows = min(100, sheet.max_row)
            max_cols = min(20, sheet.max_column)
            
            for row in range(1, max_rows + 1):
                for col in range(1, max_cols + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    if hasattr(cell, 'data_validation') and cell.data_validation:
                        validation = cell.data_validation
                        
                        # Create a key to identify this validation rule
                        rule_key = f"{validation.type}_{validation.operator}_{validation.formula1}"
                        
                        if rule_key not in processed_rules:
                            processed_rules.add(rule_key)
                            
                            # Get the cell reference
                            cell_ref = f"{get_column_letter(col)}{row}"
                            
                            # Create a token-efficient representation
                            rule_data = {
                                'cell': cell_ref,
                                'type': validation.type,
                                'formula1': validation.formula1
                            }
                            
                            # Only add non-None values for optional fields
                            if validation.operator:
                                rule_data['operator'] = validation.operator
                            
                            if validation.formula2:
                                rule_data['formula2'] = validation.formula2
                                
                            if hasattr(validation, 'showDropDown') and validation.showDropDown is not None:
                                rule_data['dropdown'] = not validation.showDropDown
                            
                            # For list type, extract the list values
                            if validation.type == 'list' and validation.formula1:
                                list_formula = validation.formula1
                                if list_formula.startswith('='):
                                    list_formula = list_formula[1:]
                                
                                # If it's a direct list like "A,B,C"
                                if list_formula.startswith('"') and list_formula.endswith('"'):
                                    try:
                                        values = list_formula.strip('"').split(',')
                                        rule_data['values'] = values
                                    except:
                                        pass
                            
                            sheet_rules.append(rule_data)
                            
                            # Limit to 10 rules per sheet for token efficiency
                            if len(sheet_rules) >= 10:
                                break
                
                if len(sheet_rules) >= 10:
                    break
            
            if sheet_rules:
                validation_rules[sheet_name] = sheet_rules
    
    except Exception as e:
        print_status(f"Warning: Could not extract data validation rules: {str(e)}", 'warning')
    
    return validation_rules

def normalize_formula(formula):
    """
    Normalize an Excel formula by replacing specific cell references with relative patterns.
    This helps identify similar formula patterns across different cells.
    
    Example: "=F15-D16+E16" becomes "=F[n]-D[n+1]+E[n+1]"
    """
    if not formula:
        return ""
    
    # Remove equals sign for processing
    if formula.startswith('='):
        formula = formula[1:]
    
    # Process absolute references (with $ signs)
    formula = formula.replace('$', '')
    
    # Use regex to find cell references
    # This pattern matches column letters followed by row numbers (A1, BC123, etc.)
    pattern = r'([A-Z]+)(\d+)'
    
    # Track the "base row" for relative positioning
    base_row = None
    refs = {}
    
    # Find all cell references
    matches = re.findall(pattern, formula)
    if not matches:
        return f"={formula}"  # Return unchanged if no cell references found
    
    # Determine base row from first reference
    for col, row in matches:
        if base_row is None:
            base_row = int(row)
            break
    
    # Replace each reference with a pattern
    normalized = formula
    for col, row in matches:
        row_num = int(row)
        offset = row_num - base_row
        
        # Create a position marker that indicates relative position
        if offset == 0:
            position = "[n]"
        elif offset > 0:
            position = f"[n+{offset}]"
        else:
            position = f"[n{offset}]"  # Negative already has - sign
        
        # Replace in the formula (careful to only replace exact matches)
        orig_ref = f"{col}{row}"
        pattern_ref = f"{col}{position}"
        normalized = re.sub(r'\b' + re.escape(orig_ref) + r'\b', pattern_ref, normalized)
    
    return f"={normalized}"

def get_formula_pattern_signature(cell):
    """
    Generate a signature for a cell's formula pattern.
    This identifies the formula structure regardless of specific cell references.
    """
    if not cell or not hasattr(cell, 'data_type') or cell.data_type != 'f':
        return None
    
    formula = cell.value
    if not formula:
        return None
    
    if isinstance(formula, str) and formula.startswith('='):
        normalized = normalize_formula(formula)
        return normalized
    
    return None

def classify_rows(sheet, options=None):
    """
    Classify each row in the sheet based on its formula patterns.
    
    Returns:
        Dict with:
            'unique_formula_rows': Set of row indices containing unique formula patterns
            'duplicate_formula_rows': Dict mapping row indices to their formula pattern signatures
            'raw_data_rows': Set of row indices containing only raw data (no formulas)
            'pattern_to_rows': Dict mapping formula patterns to sets of row indices
    """
    if options is None:
        options = {}
    
    # Initialize result structures
    classification = {
        'unique_formula_rows': set(),
        'duplicate_formula_rows': {},
        'raw_data_rows': set(),
        'pattern_to_rows': {},
        'header_rows': set(range(1, min(6, sheet.max_row + 1))),  # First 5 rows assumed to be headers
        'footer_rows': set(range(max(1, sheet.max_row - 4), sheet.max_row + 1)),  # Last 5 rows assumed to be footers
        'formula_signatures': {}  # Map row index to list of formula signatures in that row
    }
    
    # First, identify all unique formula patterns
    known_patterns = set()
    row_patterns = {}
    
    # Examine each cell in the sheet
    for row in range(1, sheet.max_row + 1):
        row_has_formula = False
        row_signatures = []
        
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            
            # Check if cell has a formula
            if cell.data_type == 'f':
                row_has_formula = True
                signature = get_formula_pattern_signature(cell)
                if signature:
                    row_signatures.append(signature)
                    
                    # Track unique patterns
                    if signature not in known_patterns:
                        known_patterns.add(signature)
                        if signature not in classification['pattern_to_rows']:
                            classification['pattern_to_rows'][signature] = set()
                        classification['pattern_to_rows'][signature].add(row)
                        classification['unique_formula_rows'].add(row)
        
        # Store signatures for this row
        if row_signatures:
            classification['formula_signatures'][row] = row_signatures
        
        # Classify the row based on formulas or lack thereof
        if not row_has_formula:
            classification['raw_data_rows'].add(row)
        elif row not in classification['unique_formula_rows']:
            # This row has formulas but they're all duplicates of existing patterns
            classification['duplicate_formula_rows'][row] = row_signatures
    
    return classification

def smart_sampling_rows(sheet, options=None):
    """
    Perform intelligent sampling of rows based on formula patterns.
    
    Args:
        sheet: The worksheet to sample
        options: Dict of options including:
            'base_sampling_rate': Base sampling rate for duplicate patterns (default: 10)
            'aggressive_sampling': Whether to use more aggressive sampling for large sheets (default: True)
    
    Returns:
        Set of row indices to keep
    """
    if options is None:
        options = {
            'base_sampling_rate': 10,
            'aggressive_sampling': True
        }
    
    # Classify rows
    classification = classify_rows(sheet, options)
    
    # Initialize set of rows to keep
    rows_to_keep = set()
    
    # Always keep header and footer rows
    rows_to_keep.update(classification['header_rows'])
    rows_to_keep.update(classification['footer_rows'])
    
    # Always keep rows with unique formula patterns
    rows_to_keep.update(classification['unique_formula_rows'])
    
    # Now handle duplicate formula patterns
    pattern_counters = {}  # Track occurrence of each pattern
    
    # Calculate adaptive sampling rate based on sheet size
    base_rate = options.get('base_sampling_rate', 10)
    sheet_size = sheet.max_row
    
    if options.get('aggressive_sampling', True):
        if sheet_size < 100:
            sampling_rate = 2  # Keep every 2nd row
        elif sheet_size < 500:
            sampling_rate = base_rate  # Keep every 10th row
        elif sheet_size < 1000:
            sampling_rate = base_rate * 2  # Keep every 20th row
        elif sheet_size < 5000:
            sampling_rate = base_rate * 5  # Keep every 50th row
        else:
            sampling_rate = base_rate * 10  # Keep every 100th row
    else:
        sampling_rate = base_rate
    
    # Process each row with duplicate formulas
    for pattern, rows in classification['pattern_to_rows'].items():
        if len(rows) <= 2:
            # If only 1-2 rows have this pattern, keep them all
            rows_to_keep.update(rows)
            continue
        
        # Calculate pattern-specific sampling rate based on frequency
        pattern_sampling_rate = sampling_rate
        if len(rows) > 100:
            pattern_sampling_rate = sampling_rate * 2
        
        # Always keep the first and last occurrence of each pattern
        rows_list = sorted(rows)
        rows_to_keep.add(rows_list[0])  # First occurrence
        rows_to_keep.add(rows_list[-1])  # Last occurrence
        
        # Sample the middle occurrences
        for i, row in enumerate(rows_list[1:-1], 1):
            if i % pattern_sampling_rate == 0:
                rows_to_keep.add(row)
    
    # Sample raw data rows at a higher sampling rate
    raw_data_sampling_rate = sampling_rate * 2  # Even more aggressive sampling for raw data
    
    # Make sure we include some raw data for context
    raw_data_rows = sorted(classification['raw_data_rows'])
    
    if raw_data_rows:
        # Always keep first and last raw data row
        rows_to_keep.add(raw_data_rows[0])
        rows_to_keep.add(raw_data_rows[-1])
        
        # Sample the rest
        for i, row in enumerate(raw_data_rows[1:-1], 1):
            if i % raw_data_sampling_rate == 0:
                rows_to_keep.add(row)
    
    # Check if we need to include more rows for context
    # If we've sampled too aggressively, add some rows back in
    if len(rows_to_keep) < min(50, sheet.max_row / 10):
        # Add more rows for context by reducing sampling rate
        for row in range(1, sheet.max_row + 1):
            if row not in rows_to_keep and row % (raw_data_sampling_rate // 2) == 0:
                rows_to_keep.add(row)
    
    return rows_to_keep

def sample_sheet_intelligently(sheet, options=None):
    """
    Apply intelligent sampling to reduce dataset size while preserving structure.
    This is the main function that will be called from convert_spreadsheet_to_json.
    
    Args:
        sheet: The worksheet to sample
        options: Dict of sampling options
    
    Returns:
        Set of row indices to keep
    """
    print_status(f"Applying intelligent formula-based sampling to sheet: {sheet.title}", 'info')
    
    if options is None:
        options = {}
    
    # Get rows to keep based on formula patterns
    rows_to_keep = smart_sampling_rows(sheet, options)
    
    original_row_count = sheet.max_row
    sampled_row_count = len(rows_to_keep)
    reduction_pct = ((original_row_count - sampled_row_count) / original_row_count) * 100 if original_row_count > 0 else 0
    
    print_status(f"Sampled {sheet.title}: {sampled_row_count}/{original_row_count} rows kept ({reduction_pct:.1f}% reduction)", 'success')
    
    return rows_to_keep

def convert_spreadsheet_to_json(file_path: str, sample_size: Optional[int] = None, 
                               formulas_only: bool = False, 
                               keep_formatting: bool = False,
                               add_context: bool = True) -> Dict[str, Any]:
    """Convert spreadsheet to JSON structure with token efficiency options and enriched context.
    
    Args:
        file_path: Path to the Excel file
        sample_size: Optional number of rows to process per sheet
        formulas_only: If True, only include cells with formulas and their dependencies
        keep_formatting: If True, include formatting information (uses more tokens)
        add_context: If True, add enriched context for formulas
    """
    print_status(f"Processing file: {file_path}", 'info')
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        # Track cells with formulas and their dependencies
        formula_cells = {}
        dependent_cells = set()
        
        # First pass to identify formulas and dependencies
        if formulas_only:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                formula_cells[sheet_name] = {}
                
                max_rows = min(sheet.max_row, sample_size) if sample_size else sheet.max_row
                
                for row in range(1, max_rows + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.data_type == 'f':
                            cell_address = f"{get_column_letter(col)}{row}"
                            formula_cells[sheet_name][cell_address] = True
                            
                            # Extract dependencies
                            deps = get_cell_dependencies(cell, sheet)
                            if deps:
                                if TOKEN_EFFICIENT and 'cr' in deps:
                                    for ref in deps['cr']:
                                        # Handle ranges
                                        if ':' in ref:
                                            start, end = ref.split(':')
                                            dependent_cells.add((sheet_name, start))
                                            dependent_cells.add((sheet_name, end))
                                        else:
                                            dependent_cells.add((sheet_name, ref))
                                elif not TOKEN_EFFICIENT and 'cell_references' in deps:
                                    for ref in deps['cell_references']:
                                        # Handle ranges
                                        if ':' in ref:
                                            start, end = ref.split(':')
                                            dependent_cells.add((sheet_name, start))
                                            dependent_cells.add((sheet_name, end))
                                        else:
                                            dependent_cells.add((sheet_name, ref))
        
        # Create the result structure
        if TOKEN_EFFICIENT:
            result = {
                map_key('file_name'): os.path.basename(file_path),
                map_key('named_ranges'): {},
                map_key('sheets'): {}
            }
        else:
            result = {
                'file_name': os.path.basename(file_path),
                'named_ranges': {},
                'sheets': {}
            }
        
        # Process sheets
        for sheet_name in workbook.sheetnames:
            print_status(f"Processing sheet: {sheet_name}", 'info')
            sheet = workbook[sheet_name]
            
            if TOKEN_EFFICIENT:
                sheet_data = {
                    map_key('metadata'): get_sheet_metadata(sheet),
                    map_key('cells'): {}
                }
            else:
                sheet_data = {
                    'metadata': get_sheet_metadata(sheet),
                    'cells': {}
                }
            
            # Determine the number of rows to process
            max_rows = min(sheet.max_row, sample_size) if sample_size else sheet.max_row
            
            # Track how many non-formula cells we've included for sampling
            data_cell_count = 0
            max_data_cells = 100  # Maximum number of data cells to include as examples
            
            # Process cells up to sample_size if specified
            for row in range(1, max_rows + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_address = f"{get_column_letter(col)}{row}"
                    
                    # Skip empty cells
                    if cell.value is None:
                        continue
                    
                    # If formulas_only, only include formula cells and their dependencies
                    if formulas_only:
                        is_formula = sheet_name in formula_cells and cell_address in formula_cells[sheet_name]
                        is_dependency = (sheet_name, cell_address) in dependent_cells
                        
                        if not is_formula and not is_dependency and data_cell_count >= max_data_cells:
                            continue
                        
                        if not is_formula and not is_dependency:
                            data_cell_count += 1
                    
                    # Get cell value
                    cell_value = get_cell_value(cell)
                    if cell_value is None:
                        continue
                    
                    # Create cell data with minimal info
                    if TOKEN_EFFICIENT:
                        cell_data = {map_key('value'): cell_value}
                        
                        # Add dependencies if it's a formula cell
                        if cell.data_type == 'f':
                            deps = get_cell_dependencies(cell, sheet)
                            if deps:
                                cell_data[map_key('dependencies')] = deps
                    else:
                        cell_data = {'value': cell_value}
                        
                        # Add dependencies if it's a formula cell
                        if cell.data_type == 'f':
                            deps = get_cell_dependencies(cell, sheet)
                            if deps:
                                cell_data['dependencies'] = deps
                    
                    # Add to sheet data
                    if TOKEN_EFFICIENT:
                        sheet_data[map_key('cells')][cell_address] = cell_data
                    else:
                        sheet_data['cells'][cell_address] = cell_data
            
            # Add sheet data to result
            if TOKEN_EFFICIENT:
                result[map_key('sheets')][sheet_name] = sheet_data
            else:
                result['sheets'][sheet_name] = sheet_data
        
        # Add enriched context if requested
        if add_context:
            # Extract table structures
            tables = extract_table_structures(workbook)
            
            # Extract column types
            column_types = extract_column_types(workbook)
            
            # Extract formula patterns
            formula_patterns = extract_formula_patterns(result)
            
            # Extract sample calculated values
            samples = extract_sample_calculated_values(workbook, formula_cells)
            
            # Extract data validation rules
            validation_rules = extract_data_validation_rules(workbook)
            
            # Extract pivot tables
            pivot_tables = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                pt_data = get_pivot_tables(sheet)
                if pt_data:
                    pivot_tables[sheet_name] = pt_data
            
            # Get implementation notes
            implementation_notes = get_implementation_notes(formula_patterns)
            
            # Add to result
            if TOKEN_EFFICIENT:
                enriched_context = {
                    map_key('tables'): tables,
                    map_key('column_types'): column_types,
                    map_key('formula_patterns'): formula_patterns,
                    map_key('implementation_notes'): implementation_notes
                }
                
                # Only add samples, validation and pivot tables if they exist (to save tokens)
                if samples:
                    enriched_context[map_key('samples')] = samples
                    
                if validation_rules:
                    enriched_context[map_key('validation')] = validation_rules
                    
                if pivot_tables:
                    enriched_context[map_key('pivot_tables')] = pivot_tables
                    
                result[map_key('enriched_context')] = enriched_context
            else:
                enriched_context = {
                    'tables': tables,
                    'column_types': column_types,
                    'formula_patterns': formula_patterns,
                    'implementation_notes': implementation_notes
                }
                
                # Only add samples, validation and pivot tables if they exist (to save tokens)
                if samples:
                    enriched_context['samples'] = samples
                    
                if validation_rules:
                    enriched_context['validation'] = validation_rules
                    
                if pivot_tables:
                    enriched_context['pivot_tables'] = pivot_tables
                    
                result['enriched_context'] = enriched_context
        
        return result
    
    except Exception as e:
        print_status(f"Error converting spreadsheet: {str(e)}", 'error')
        raise

def count_json_tokens(data: Dict[str, Any]) -> int:
    """
    Count the number of tokens in JSON data.
    This is an approximation that tries to match typical LLM tokenization patterns.
    """
    try:
        # Convert to JSON string using the custom encoder
        json_str = json.dumps(data, cls=ExcelJSONEncoder)
        
        # Basic tokenization rules that approximate LLM tokenizers:
        # 1. Split on whitespace
        # 2. Split on punctuation
        # 3. Split on case changes (camelCase, PascalCase)
        # 4. Keep common JSON structural elements as single tokens
        
        # First, handle JSON structural elements
        json_str = re.sub(r'([\{\}\[\],:])', r' \1 ', json_str)
        
        # Split on camelCase and PascalCase
        json_str = re.sub(r'([a-z])([A-Z])', r'\1 \2', json_str)
        json_str = re.sub(r'([A-Z])([A-Z][a-z])', r'\1 \2', json_str)
        
        # Handle numbers and special characters
        json_str = re.sub(r'([0-9]+)', r' \1 ', json_str)
        json_str = re.sub(r'([^a-zA-Z0-9\s\{\}\[\],:])', r' \1 ', json_str)
        
        # Split and filter empty strings
        tokens = [token for token in json_str.split() if token.strip()]
        
        # Add a small overhead for JSON structure and formatting
        structural_overhead = len(re.findall(r'[\{\}\[\],:]', json_str)) * 0.5
        
        # Return total token count with overhead
        return int(len(tokens) + structural_overhead)
        
    except Exception as e:
        print_status(f"Error counting tokens: {str(e)}", 'error')
        return 0

def save_json_output(data: Dict[str, Any], original_file_path: str, minify: bool = False) -> str:
    """Save JSON output to file with token count in filename."""
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Count tokens
        token_count = count_json_tokens(data)
        print_status(f"Total tokens in output: {token_count}", 'info')
        
        # Generate filename with token count and timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(original_file_path))[0]
        
        # Indicate if token-efficient mode was used
        mode_suffix = "_efficient" if TOKEN_EFFICIENT else ""
        
        output_filename = f"{base_name}{mode_suffix}_{token_count}tokens_{timestamp}.json"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        # Add metadata including token count
        metadata_key = map_key('metadata') if TOKEN_EFFICIENT else 'metadata'
        if metadata_key not in data:
            data[metadata_key] = {}
        
        if TOKEN_EFFICIENT:
            data[metadata_key].update({
                map_key('token_count'): token_count,
                map_key('conversion_timestamp'): timestamp,
                map_key('original_filename'): os.path.basename(original_file_path)
            })
        else:
            data[metadata_key].update({
                'token_count': token_count,
                'conversion_timestamp': timestamp,
                'original_filename': os.path.basename(original_file_path)
            })
        
        # Save the file using the custom encoder (with or without indentation)
        with open(output_path, 'w', encoding=DEFAULT_ENCODING) as f:
            if minify:
                json.dump(data, f, ensure_ascii=False, cls=ExcelJSONEncoder, separators=(',', ':'))
            else:
                json.dump(data, f, indent=2, ensure_ascii=False, cls=ExcelJSONEncoder)
        
        print_status(f"Successfully saved JSON to: {output_path} ({token_count} tokens)", 'success')
        return output_path
    
    except Exception as e:
        print_status(f"Error saving JSON output: {str(e)}", 'error')
        raise

def convert_spreadsheet_to_json_with_sampling(file_path: str, sample_size: Optional[int] = None, 
                               formulas_only: bool = False, 
                               keep_formatting: bool = False,
                               add_context: bool = True,
                               intelligent_sampling: bool = True,
                               sampling_options: Optional[Dict] = None) -> Dict[str, Any]:
    """Convert spreadsheet to JSON structure with intelligent formula-preserving sampling.
    
    Args:
        file_path: Path to the Excel file
        sample_size: Optional number of rows to process per sheet
        formulas_only: If True, only include cells with formulas and their dependencies
        keep_formatting: If True, include formatting information (uses more tokens)
        add_context: If True, add enriched context for formulas
        intelligent_sampling: If True, use formula-preserving intelligent sampling
        sampling_options: Options for intelligent sampling
    """
    if sampling_options is None:
        sampling_options = {
            'base_sampling_rate': 10,
            'aggressive_sampling': True
        }
        
    print_status(f"Processing file with intelligent sampling: {file_path}", 'info')
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        # Track cells with formulas and their dependencies
        formula_cells = {}
        dependent_cells = set()
        
        # Create the result structure
        if TOKEN_EFFICIENT:
            result = {
                map_key('file_name'): os.path.basename(file_path),
                map_key('named_ranges'): {},
                map_key('sheets'): {}
            }
        else:
            result = {
                'file_name': os.path.basename(file_path),
                'named_ranges': {},
                'sheets': {}
            }
        
        # Process sheets
        for sheet_name in workbook.sheetnames:
            print_status(f"Processing sheet: {sheet_name}", 'info')
            sheet = workbook[sheet_name]
            
            if TOKEN_EFFICIENT:
                sheet_data = {
                    map_key('metadata'): get_sheet_metadata(sheet),
                    map_key('cells'): {}
                }
            else:
                sheet_data = {
                    'metadata': get_sheet_metadata(sheet),
                    'cells': {}
                }
            
            # Apply intelligent sampling if enabled
            if intelligent_sampling and sheet.max_row > 50:  # Only sample larger sheets
                rows_to_keep = sample_sheet_intelligently(sheet, sampling_options)
            else:
                # Without intelligent sampling, keep all rows up to sample_size
                rows_to_keep = set(range(1, min(sheet.max_row + 1, (sample_size or sheet.max_row) + 1)))
            
            # Process cells from the kept rows
            for row in rows_to_keep:
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_address = f"{get_column_letter(col)}{row}"
                    
                    # Skip empty cells
                    if cell.value is None:
                        continue
                    
                    # Get cell value
                    cell_value = get_cell_value(cell)
                    if cell_value is None:
                        continue
                    
                    # Create cell data with minimal info
                    if TOKEN_EFFICIENT:
                        cell_data = {map_key('value'): cell_value}
                        
                        # Add dependencies if it's a formula cell
                        if cell.data_type == 'f':
                            deps = get_cell_dependencies(cell, sheet)
                            if deps:
                                cell_data[map_key('dependencies')] = deps
                    else:
                        cell_data = {'value': cell_value}
                        
                        # Add dependencies if it's a formula cell
                        if cell.data_type == 'f':
                            deps = get_cell_dependencies(cell, sheet)
                            if deps:
                                cell_data['dependencies'] = deps
                    
                    # Add to sheet data
                    if TOKEN_EFFICIENT:
                        sheet_data[map_key('cells')][cell_address] = cell_data
                    else:
                        sheet_data['cells'][cell_address] = cell_data
            
            # Add sheet data to result
            if TOKEN_EFFICIENT:
                result[map_key('sheets')][sheet_name] = sheet_data
            else:
                result['sheets'][sheet_name] = sheet_data
        
        # Add named ranges
        named_ranges = get_named_ranges(workbook)
        if TOKEN_EFFICIENT:
            result[map_key('named_ranges')] = named_ranges
        else:
            result['named_ranges'] = named_ranges
        
        # Add enriched context if requested
        if add_context:
            # Extract table structures, column types, formula patterns, etc.
            tables = extract_table_structures(workbook)
            column_types = extract_column_types(workbook)
            formula_patterns = extract_formula_patterns(result)
            implementation_notes = get_implementation_notes(formula_patterns)
            
            # Add to result
            if TOKEN_EFFICIENT:
                enriched_context = {
                    map_key('tables'): tables,
                    map_key('column_types'): column_types,
                    map_key('formula_patterns'): formula_patterns,
                    map_key('implementation_notes'): implementation_notes
                }
                result[map_key('enriched_context')] = enriched_context
            else:
                enriched_context = {
                    'tables': tables,
                    'column_types': column_types,
                    'formula_patterns': formula_patterns,
                    'implementation_notes': implementation_notes
                }
                result['enriched_context'] = enriched_context
        
        return result
    
    except Exception as e:
        print_status(f"Error converting spreadsheet: {str(e)}", 'error')
        raise

def main():
    if len(sys.argv) < 2:
        print_status("Usage: python spreadsheet_converter.py <path_to_spreadsheet> [sample_size] [--formulas-only] [--keep-formatting] [--minify] [--full-names] [--no-context] [--intelligent-sampling]", 'error')
        print_status("  sample_size: Optional. Number of rows to process per sheet", 'info')
        print_status("  --formulas-only: Only include cells with formulas and their dependencies", 'info')
        print_status("  --keep-formatting: Include formatting information (uses more tokens)", 'info')
        print_status("  --minify: Remove whitespace from JSON output", 'info')
        print_status("  --full-names: Use full property names instead of abbreviated ones", 'info')
        print_status("  --no-context: Skip adding enriched context information", 'info')
        print_status("  --intelligent-sampling: Use formula-preserving intelligent sampling", 'info')
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    # Parse arguments
    sample_size = None
    formulas_only = '--formulas-only' in sys.argv
    keep_formatting = '--keep-formatting' in sys.argv
    minify = '--minify' in sys.argv
    no_context = '--no-context' in sys.argv
    intelligent_sampling = '--intelligent-sampling' in sys.argv
    
    # Set token efficiency based on arguments
    global TOKEN_EFFICIENT
    TOKEN_EFFICIENT = '--full-names' not in sys.argv
    
    # Check for sample size
    for arg in sys.argv[2:]:
        if arg.isdigit():
            sample_size = int(arg)
            if sample_size <= 0:
                print_status("Sample size must be a positive integer", 'error')
                sys.exit(1)
    
    if not os.path.exists(file_path):
        print_status(f"File not found: {file_path}", 'error')
        sys.exit(1)
    
    if not any(file_path.lower().endswith(ext) for ext in SUPPORTED_EXTENSIONS):
        print_status(f"Unsupported file format. Supported formats: {', '.join(SUPPORTED_EXTENSIONS)}", 'error')
        sys.exit(1)
    
    try:
        # Use the new intelligent sampling function if enabled
        if intelligent_sampling:
            data = convert_spreadsheet_to_json_with_sampling(
                file_path, 
                sample_size=sample_size,
                formulas_only=formulas_only,
                keep_formatting=keep_formatting,
                add_context=not no_context,
                intelligent_sampling=True
            )
        else:
            # Use the original conversion function
            data = convert_spreadsheet_to_json(
                file_path, 
                sample_size=sample_size,
                formulas_only=formulas_only,
                keep_formatting=keep_formatting,
                add_context=not no_context
            )
        
        # Save the result
        save_json_output(data, file_path, minify)
        
    except Exception as e:
        print_status(f"Error: {str(e)}", 'error')
        sys.exit(1)

if __name__ == "__main__":
    main() 